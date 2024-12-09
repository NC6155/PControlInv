from django.shortcuts import render, redirect, get_object_or_404
from .models import Tables
from django.core.paginator import Paginator
from .forms import *
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib import messages



# Create your views here.


def tables(request):
    tablesEntries=Tables.objects.all().order_by('id')
    form=TablesSearchForm(request.POST or None)
    paginator = Paginator(tablesEntries, 5) #paginador, dividido por 5 entradas en cada pagina
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    context = {
        "form": form,
        "everyEntry": tablesEntries,
        "page_obj": page_obj,
        } #contexto para la página, para definirlo de una sola forma
    if request.method == 'POST': #Si hubo un método post dentro del html
        page_obj = Tables.objects.filter(codProd__icontains=form['codProd'].value(), #filtra las tablas por estos valores
                                        nomProd__icontains=form['nomProd'].value()
                                        )
        context = {
        "form": form,
        "everyEntry": tablesEntries,
        "page_obj": page_obj,
        }
    return render(request, "stockList/tables.html", context)

def add_stock(request):
    form = TablesCreateForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Guardado exitosamente')
        return redirect('/adding_stock/')
    context = {
		"form": form,
		"title": "Add Item",
	}
    return render(request, "stockList/add_stock.html", context)

def update_stock(request, pk):
    queryset = Tables.objects.get(id=pk)
    form = TablesUpdateForm(instance=queryset)
    if request.method == 'POST':
        form = TablesUpdateForm(request.POST, instance=queryset)
        if form.is_valid():
            form.save()
            messages.success(request, 'Actualizado exitosamente')
            return redirect('/tables/')
    context = {
		'form':form
    }
    return render(request, 'stockList/add_stock.html', context)

def delete_stock(request, pk):
    queryset = Tables.objects.get(id=pk)
    if request.method == 'POST':
        queryset.delete()
        messages.success(request, 'Eliminado exitosamente')
        return redirect('/tables/')
    return render(request, 'stockList/delete_stock.html')

def reducir_stock(request):
    mensaje = ""
    if request.method == 'POST':
        codProd = request.POST.get('codProd')
        cantidad = int(request.POST.get('cantidad'))

        # Buscar todos los productos por su código
        productos = Tables.objects.filter(codProd=codProd)

        if productos.exists():
            for producto in productos:
                if cantidad > 0 and producto.stock >= cantidad:
                    # Reducir el stock
                    producto.stock -= cantidad
                    producto.save()
                    mensaje = f'Stock del producto {producto.nomProd} reducido en {cantidad}. Stock actual: {producto.stock}.'
                else:
                    mensaje = 'Cantidad no válida o stock insuficiente.'
        else:
            mensaje = 'Producto no encontrado.'

    context = {
        'mensaje': mensaje
    }

    return render(request, 'stockList/reducir_stock.html', context)



class ReporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        productos = Tables.objects.all()
        wb = Workbook()
        ws = wb.active
        
        
        ws['B1'] = 'Productos de Bodega'
        ws.merge_cells('B1:E1')
        ws['B1'].font = Font(size=14, bold=True)
        ws['B1'].alignment = Alignment(horizontal='center')
        
        
        headers = ['Código Producto', 'Nombre Producto', 'Calificación', 'Tipo de Producto', 'Stock']
        col_idx = 2
        for header in headers:
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            col_idx += 1
        
      
        cont = 5
        for producto in productos:
            ws.cell(row=cont, column=2).value = producto.codProd
            ws.cell(row=cont, column=3).value = producto.nomProd
            ws.cell(row=cont, column=4).value = producto.calificacion
            ws.cell(row=cont, column=5).value = producto.tipoProd
            ws.cell(row=cont, column=6).value = producto.stock
            cont += 1
        
        
        for col in ws.iter_cols(min_col=2, max_col=6):
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        nombre_archivo = "Productos_Bodega.xlsx"
        response = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

